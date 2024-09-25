import numpy as np
import pandas as pd
np.set_printoptions(suppress=True, precision=2)
pd.options.display.float_format = '{:,.2f}'.format

import openpyxl as op
import duckdb

import os
import sys
from datetime import datetime


try:
    LOCAL_DIR = os.path.dirname(__file__)
except:
    LOCAL_DIR = os.getcwd()

PARENT_DIR = os.path.dirname(LOCAL_DIR)
sys.path.append(PARENT_DIR)

import tbl_calc

REPORT_MTH = tbl_calc.REPORT_MTH
DB_FILE = tbl_calc.DB_FILE

print('\n5. Средняя линия')
print('-'*79)


# Словарь для сортировки значений категорий товаров, форматов магазинов и производителей
rename_dict = {
    'group': {
        'Группа 1': 'a. Группа 1',
        'Группа 2': 'b. Группа 2',
        'Группа 3': 'c. Группа 3',
        'Группа 4': 'd. Группа 4',
        'Группа 5': 'e. Группа 5',
        'Группа 6': 'f. Группа 6'
    },
    'store_format': {
        'Формат 1': 'a. Формат 1', 
        'Формат 2': 'b. Формат 2',
        'Формат 3': 'c. Формат 3', 
        'Формат 4': 'd. Формат 4'
    },
    'producer': {
        'Производитель 1':  'a. Производитель 1',
        'Производитель 2':  'b. Производитель 2',
        'Производитель 3':  'c. Производитель 3',
        'Производитель 4':  'd. Производитель 4',
        'Производитель 5':  'e. Производитель 5',
        'Производитель 6':  'f. Производитель 6',
        'Производитель 7':  'g. Производитель 7',
        'Производитель 8':  'h. Производитель 8',
        'Производитель 9':  'i. Производитель 9',
        'Производитель 10': 'j. Производитель 10',
        'Производитель 11': 'k. Производитель 11',
        'Производитель 12': 'l. Производитель 12'
    }
}


# Списки для итерации
groups = list(rename_dict['group'].values())
store_formats = list(rename_dict['store_format'].values())
producers = list(rename_dict['producer'].values())


try:
    conx = duckdb.connect(database=DB_FILE, read_only=True)
except Exception as err:
    print(f'Не удается подключиться к базе данных:\n{err}')


try:
    with (open(os.path.join(LOCAL_DIR, '5. Средняя линия.sql'), encoding='utf-8') as q1, 
          open(os.path.join(LOCAL_DIR, '5. Средняя линия (ХХ).sql'), encoding='utf-8') as q2):
        sql_queries_1, sql_queries_2 = q1.read().split('--split'), q2.read().split('--split')
except Exception as err:
    print(f'Не удается открыть файл с SQL запросом:\n{err}')


dfs_names_in = ['grp', 'grp_prod', 'grp_fmt', 'grp_fmt_prod']
grp, grp_prod, grp_fmt, grp_fmt_prod = None, None, None, None


for df, query_1, query_2 in zip(dfs_names_in, sql_queries_1, sql_queries_2):
    globals()[df] = pd.concat([
        conx.execute(query_1).fetch_df(),
        conx.execute(query_2).fetch_df()
    ])


conx.close()


writer = pd.ExcelWriter(os.path.join(LOCAL_DIR, 'data.xlsx'))

for df in dfs_names_in:
    globals()[df]['sl'] = globals()[df]['skus']/globals()[df]['akb']
    globals()[df].to_excel(writer, sheet_name=f'{df}', index=False)

writer.close()


# Создание словаря c названиями листов и датафреймами
dfs = pd.read_excel(os.path.join(LOCAL_DIR, 'data.xlsx'), sheet_name=None)

# Наименования а) четырех датафреймов с данными из Excel файла, б) результирующих датафреймов и с) листов выходного Excel файла
dfs_names_in = ['grp', 'grp_prod', 'grp_fmt', 'grp_fmt_prod']
dfs_names_out = ['df_' + str(i) for i in range(7)]
ws_names = ['Сети', 'Группа 1', 'Группа 2', 'Группа 3', 'Группа 4', 'Группа 5', 'Группа 6']

# Проверка на правильное кол-во листов в Excel файле
if len(dfs_names_in) != len(dfs):
    raise Exception(f'Expected 4 worksheets in Excel file. Got {len(dfs)} worksheets.')

# Присвоение пустых значений четырем переменным с исходными данными
grp, grp_prod, grp_fmt, grp_fmt_prod = None, None, None, None

for i, sheet in enumerate(dfs.values()):
    globals()[dfs_names_in[i]] = sheet

# Пустые результирующие датафреймы
for df in dfs_names_out:
    globals()[df] = pd.DataFrame()


# Для каждого датафрейма:
# - унифицировать кол-во столбцов
# - заполнить пустые значения словом 'Все'
# - сформировать сводную таблицу
# - отсортировать строки используя пользовательский порядок
# - отсортировать индекс стролбцов (месяцы)
# - заменить формат дат с Timestamp на Date, на строку
# - перенести индекс строк в столбцы

for df in dfs_names_in:
    globals()[df] = (
        globals()[df].reindex(columns=grp_fmt_prod.columns)
        .fillna('Все')
        .replace(rename_dict)
        .pivot_table(
            index=['group', 'store_format', 'producer'],
            columns='month_year', 
            values='sl',
            aggfunc='max',
            fill_value=0,
            dropna=False,
            sort=False
        )
        .sort_values(
            axis='index',
            by=['group', 'store_format', 'producer'], 
            ascending=[True, True, True]
        )
        .sort_index(axis='columns')
    )
    globals()[df].columns = [d.date() for d in globals()[df].columns]
    globals()[df].columns = [col.strftime('%d.%m.%Y') for col in globals()[df].columns]
    globals()[df] = globals()[df].reset_index()


# Первый результирующий датафрейм с данными по категориям и производителям
for group in groups:
    df_0 = pd.concat([df_0, grp[grp['group'] == group]])
    df_0 = pd.concat([df_0, grp_prod[grp_prod['group'] == group]])


# Добавить единый столбец с описанием, удалить ненужные столбцы и убрать информацию для сортировки
df_0.insert(
    0, 
    'title', 
    df_0.apply(
        lambda row: row['group'] if row['producer'] == 'Все' else row['producer'] , 
        axis='columns'
        )
)

df_0 = (
    df_0.drop(columns=list(rename_dict.keys()))
    .replace(r'^\w\S\s', '', regex=True)
)


# Второй-седьмой результирующие датафреймы с данными по категориям, форматам магазинов и производителям
for i, group in enumerate(groups):
    for store_format in store_formats:
        globals()[dfs_names_out[i+1]] = pd.concat([
                globals()[dfs_names_out[i+1]],
                grp_fmt[(grp_fmt['group'] == group) & (grp_fmt['store_format'] == store_format)]
            ])
        globals()[dfs_names_out[i+1]] = pd.concat([
            globals()[dfs_names_out[i+1]],
            grp_fmt_prod[(grp_fmt_prod['group'] == group) & (grp_fmt_prod['store_format'] == store_format)]
            ])


# Добавить единый столбец с описанием, удалить ненужные столбцы и убрать информацию для сортировки
for df in dfs_names_out[1:]:
    globals()[df].insert(
    0, 
    'title', 
    globals()[df].apply(
        lambda row: row['store_format'] if row['producer'] == 'Все' else row['producer'] , 
        axis='columns'
        )
    )

    globals()[df] = (
        globals()[df].drop(columns=list(rename_dict.keys()))
        .replace(r'^\w\S\s', '', regex=True)
    )


# Удалить строки с нулевыми значениями в каждом из рассматриваемых периодов
for df in dfs_names_out:
    globals()[df] = globals()[df].loc[~(globals()[df].iloc[:, 1:] == 0).all(axis=1), :]


# Добавить вычисление показателей
for df in dfs_names_out:
    globals()[df] = (pd.concat([
                globals()[df].set_index('title'), 
                globals()[df].set_index('title').apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m_last_year', calc_method='pp_change', aggfunc='sum').rename('m_on_m_ly_abs')
            ], axis=1)
            .pipe(tbl_calc.insert_wsp, 1)
            .reset_index(names='title')
        )
    

# Сохранить результирующие датафреймы на отдельных листах выходного Excel файла
writer = pd.ExcelWriter(os.path.join(LOCAL_DIR, 'output.xlsx'))

for i, df in enumerate(dfs_names_out):
    globals()[df].to_excel(writer, sheet_name=f'{ws_names[i]}', index=False)

writer.close()


# Базовое форматирование
col_widths = [25] + [12]*(df_0.shape[1] - 1)

wb = op.load_workbook(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))
side = op.styles.Side(border_style=None)
no_border = op.styles.borders.Border(left=side, right=side, top=side, bottom=side)
font_txt = op.styles.Font(name='Arial', size=10)
font_head = op.styles.Font(name='Arial', size=10, bold=True)

for ws in wb:
    for col in ws.iter_cols():
        ws.column_dimensions[op.utils.get_column_letter(col[0].column)].width = col_widths[col[0].column - 1]
        for cell in col:
            if cell.column > 1 and cell.row > 1:
                cell.number_format = '#,##0.0'
            if cell.row == 1:
                cell.font = font_head
                cell.border = no_border
            else:
                cell.font = font_txt
            if cell.column == 1 and (cell.value in [x[3:] for x in groups] or cell.value in [x[3:] for x in store_formats]):
                cell.font = font_head

wb.save(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))


print(f'{datetime.now().strftime("%H:%M:%S")}> Готово')