import numpy as np
import pandas as pd
np.set_printoptions(suppress=True, precision=2)
pd.options.display.float_format = '{:,.2f}'.format

import openpyxl as op
import duckdb

import os
import sys
from datetime import datetime


try:
    LOCAL_DIR = os.path.dirname(__file__)
except:
    LOCAL_DIR = os.getcwd()

PARENT_DIR = os.path.dirname(LOCAL_DIR)
sys.path.append(PARENT_DIR)

import tbl_calc

REPORT_MTH = tbl_calc.REPORT_MTH
DB_FILE = tbl_calc.DB_FILE


print('\n9. Доля НК')
print('-'*79)


producers_dict = {
    'Все производители':    'a. Все производители',
    'XX':                   'b. XX',
    'Другие производители': 'c. Другие производители'
}
groups_dict = {
    'Группа 1': 'a. Группа 1',
    'Группа 2': 'b. Группа 2',
    'Группа 3': 'c. Группа 3'
}
chains_dict = {
    'Сеть 1': 'd. Сеть 1',
    'Сеть 2': 'e. Сеть 2',
    'Сеть 3': 'f. Сеть 3',
    'Сеть 4': 'g. Сеть 4',
    'Сеть 5': 'h. Сеть 5',
    'Сеть 6': 'i. Сеть 6',
    'Сеть 7': 'j. Сеть 7'
}


try:
    conx = duckdb.connect(database=DB_FILE, read_only=True)
except Exception as err:
    print(f'Не удается подключиться к базе данных:\n{err}')


try:
    with open(os.path.join(LOCAL_DIR, '9. Доля XX.sql'), encoding='utf-8') as q:
        sql_query = q.read()
except Exception as err:
    print(f'Не удается открыть файл с SQL запросом:\n{err}')


df = conx.execute(sql_query).fetch_df()
conx.close()


df.to_csv(os.path.join(LOCAL_DIR, 'data.csv'), index=False)


df['network_subname'] = df['network_subname'].replace(chains_dict)
df['group'] = df['group'].replace(groups_dict)
df['producer'] = df['producer'].replace(producers_dict)


df.insert(3, 'net_group', '--')


idx = df.query("`network_subname`.isna() & `group`.isna()").index
df.loc[idx, 'net_group'] = 'Итого'

idx = df.query("`network_subname`.isna() & `group`.notna()").index
df.loc[idx, 'net_group'] = df.loc[idx, 'group']

idx = df.query("`network_subname`.notna()").index
df.loc[idx, 'net_group'] = df.loc[idx, 'network_subname']


df = df.fillna('[NULL]')
df = df.drop(columns='network_subname')
df = df.pivot_table(index=['group', 'net_group', 'producer'], columns='report_date', values='tn')
df = df.reset_index()
df[['group', 'net_group', 'producer']] = df[['group', 'net_group', 'producer']].replace(r'^\w\S\s', '', regex=True)
df = df.set_index(['group', 'net_group', 'producer'])


df.to_excel(os.path.join(LOCAL_DIR, 'output.xlsx'), index=True)


# Базовое форматирование
col_widths = [18, 18, 24, 12, 12]

wb = op.load_workbook(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))
side = op.styles.Side(border_style=None)
no_border = op.styles.borders.Border(left=side, right=side, top=side, bottom=side)
font_txt = op.styles.Font(name='Arial', size=10)
font_head = op.styles.Font(name='Arial', size=10, bold=True)
for ws in wb:
    for col in ws.iter_cols():
        ws.column_dimensions[op.utils.get_column_letter(col[0].column)].width = col_widths[col[0].column - 1]
        for cell in col:
            if cell.column >= 3:
                cell.number_format = '#,##0'
            if cell.row == 1 or cell.column <= 3:
                cell.font = font_head
                cell.border = no_border
            else:
                cell.font = font_txt
wb.save(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))


print(f'{datetime.now().strftime("%H:%M:%S")}> Готово')