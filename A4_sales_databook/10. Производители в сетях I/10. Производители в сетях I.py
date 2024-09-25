import numpy as np
import pandas as pd
np.set_printoptions(suppress=True, precision=2)
pd.options.display.float_format = '{:,.2f}'.format

import openpyxl as op
import duckdb

import os
import sys
from datetime import datetime


try:
    LOCAL_DIR = os.path.dirname(__file__)
except:
    LOCAL_DIR = os.getcwd()

PARENT_DIR = os.path.dirname(LOCAL_DIR)
sys.path.append(PARENT_DIR)

import tbl_calc

REPORT_MTH = tbl_calc.REPORT_MTH
DB_FILE = tbl_calc.DB_FILE


print('\n10. Производители в сетях I')
print('-'*79)


network_list = ['network_1', 'network_2', 'network_3', 'network_4', 'network_5', 'network_6', 'network_7', 'network_8', 'network_9']
groups_dict = {
    'Группа 1': 'a. Группа 1', 
    'Группа 2': 'b. Группа 2', 
    'Группа 3': 'c. Группа 3',
    'Группа 4': 'd. Группа 4'
}


try:
    conx = duckdb.connect(database=DB_FILE, read_only=True)
except Exception as err:
    print(f'Не удается подключиться к базе данных:\n{err}')


try:
    with open(os.path.join(LOCAL_DIR, '10. Производители в сетях I.sql'), encoding='utf-8') as q:
        sql_query = q.read()
except Exception as err:
    print(f'Не удается открыть файл с SQL запросом:\n{err}')


df = conx.execute(sql_query).fetch_df()
conx.close()


df.to_csv(os.path.join(LOCAL_DIR, 'data.csv'), index=False)


df['group'] = df['group'].replace(groups_dict)
grouped = df.groupby("group")


writer = pd.ExcelWriter(os.path.join(LOCAL_DIR, 'output.xlsx'))

for name, group in grouped:
    group_name = name
    df_tmp = group

    df_sort_key = df_tmp[['producer', 'tn_sort']]
    df_sort_key = df_sort_key.groupby('producer', as_index=False).sum()

    df_tmp = df_tmp.drop(['tn_curr', 'tn_prev', 'tn_sort'], axis=1)
    df_tmp = df_tmp.pivot(index='producer', columns='network_subname', values='tn_change')
    df_tmp = df_tmp.merge(df_sort_key, on='producer')
    df_tmp = df_tmp.sort_values(by='tn_sort', ascending=False)
    df_tmp = df_tmp.set_index('producer')
    df_tmp = df_tmp.reindex(columns=network_list)
    df_tmp = df_tmp.reset_index()

    df_tmp.to_excel(writer, sheet_name=name[3:])

writer.close()


# Базовое форматирование
col_widths = [8, 30] + [12]*9

wb = op.load_workbook(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))
side = op.styles.Side(border_style=None)
no_border = op.styles.borders.Border(left=side, right=side, top=side, bottom=side)
font_txt = op.styles.Font(name='Arial', size=10)
font_head = op.styles.Font(name='Arial', size=10, bold=True)
for ws in wb:
    for col in ws.iter_cols():
        ws.column_dimensions[op.utils.get_column_letter(col[0].column)].width = col_widths[col[0].column - 1]
        for cell in col:
            if cell.row == 1 or cell.column == 1:
                cell.font = font_head
                cell.border = no_border
            else:
                cell.font = font_txt
wb.save(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))


print(f'{datetime.now().strftime("%H:%M:%S")}> Готово')