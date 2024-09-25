import numpy as np
import pandas as pd
np.set_printoptions(suppress=True, precision=2)
pd.options.display.float_format = '{:,.2f}'.format

import openpyxl as op
import duckdb

import os
import sys
from datetime import datetime


try:
    LOCAL_DIR = os.path.dirname(__file__)
except:
    LOCAL_DIR = os.getcwd()

PARENT_DIR = os.path.dirname(LOCAL_DIR)
sys.path.append(PARENT_DIR)

import tbl_calc

REPORT_MTH = tbl_calc.REPORT_MTH
DB_FILE = tbl_calc.DB_FILE


print('\n14. Представленность SKU по форматам')
print('-'*79)


formats_dict = {
    'Формат 1': 'a. Формат 1',
    'Формат 2': 'b. Формат 2',
    'Формат 3': 'c. Формат 3',
    'Формат 4': 'd. Формат 4'
}
groups_dict = {
    'Группа 1': 'a. Группа 1',
    'Группа 2': 'b. Группа 2',
    'Группа 3': 'c. Группа 3',
    'Группа 4': 'd. Группа 4',
    'Группа 5': 'e. Группа 5',
    'Группа 6': 'f. Группа 6'
}


try:
    conx = duckdb.connect(database=DB_FILE, read_only=True)
except Exception as err:
    print(f'Не удается подключиться к базе данных:\n{err}')


try:
    with open(os.path.join(LOCAL_DIR, '14. Представленность SKU по форматам.sql'), encoding='utf-8') as q:
        sql_query = q.read()
except Exception as err:
    print(f'Не удается открыть файл с SQL запросом:\n{err}')


df = conx.execute(sql_query).fetch_df()
conx.close()


df.to_csv(os.path.join(LOCAL_DIR, 'data.csv'), index=False)


df_all = df.query("`store_format` in ['Формат 1', Формат 2', 'Формат 3', 'Формат 4']").copy()
df_all['store_format'] = df_all['store_format'].replace(formats_dict)
df_all['group'] = df_all['group'].replace(groups_dict)

df_xx = df_all[df_all['producer'] == 'XXXX'].copy()


df_all_sku = pd.pivot_table(df_all, columns='store_format', index='group', values='global_sku_name', aggfunc='count')
df_xx_sku = pd.pivot_table(df_xx, columns='store_format', index='group', values='global_sku_name', aggfunc='count')

df_all_kg = pd.pivot_table(df_all, columns='store_format', index='group', values='kg_sku', aggfunc='sum')
df_xx_kg = pd.pivot_table(df_xx, columns='store_format', index='group', values='kg_sku', aggfunc='sum')


sku = np.empty((df_all_sku.values.shape[0], df_all_sku.values.shape[1]*2))
sku[:, 0::2] = df_all_sku.values
sku[:, 1::2] = df_xx_sku.values
sku = pd.DataFrame(sku)
sku.index = [i[3:] for i in df_all_sku.index]
sku.columns = [j[3:] if i % 2 == 0 else 'в т.ч. XX' for i, j in enumerate(2*[i for i in df_all_sku.columns])]
sku = tbl_calc.margins(sku)

kg = np.empty((df_all_kg.values.shape[0], df_all_kg.values.shape[1]*2))
kg[:, 0::2] = df_all_kg.values
kg[:, 1::2] = df_xx_kg.values
kg = pd.DataFrame(kg)
kg.index = [i[3:] for i in df_all_kg.index]
kg.columns = [j[3:] if i % 2 == 0 else 'в т.ч. XX' for i, j in enumerate(2*[i for i in df_all_kg.columns])]
kg = tbl_calc.margins(kg)

kg_sku = kg/sku


sku_share = tbl_calc.margins(df_xx_sku)/tbl_calc.margins(df_all_sku)
sku_share.columns = [i[3:] for i in sku_share.columns]
sku_share = sku_share.reset_index(names='store_format').replace(r'^\w\S\s', '', regex=True)

kg_share = tbl_calc.margins(df_xx_kg)/tbl_calc.margins(df_all_kg)
kg_share.columns = [i[3:] for i in kg_share.columns]
kg_share = kg_share.reset_index(names='store_format').replace(r'^\w\S\s', '', regex=True)


kg_sku_share = pd.DataFrame(kg_sku.values[:, 1::2] / kg_sku.values[:, 0::2])
kg_sku_share.columns = [i[3:] for i in df_all_sku]
kg_sku_share.insert(0, 'store_format', '')
kg_sku_share['store_format'] = sku_share['store_format']


sku = sku.reset_index(names='store_format')
sku.insert(0, 'measure', 'SKU')

kg = kg.reset_index(names='store_format')
kg.insert(0, 'measure', 'кг')

kg_sku = kg_sku.reset_index(names='store_format')
kg_sku.insert(0, 'measure', 'кг/SKU')


sku_share.insert(0, 'measure', '% XX от SKU')
kg_share.insert(0, 'measure', '% XX от кг')
kg_sku_share.insert(0, 'measure', 'XX/Категория')


writer = pd.ExcelWriter(os.path.join(LOCAL_DIR, 'output.xlsx'))

pd.concat([sku, kg, kg_sku]).to_excel(writer, sheet_name='main', index=False)
pd.concat([sku_share, kg_share, kg_sku_share]).to_excel(writer, sheet_name='shares', index=False)


writer.close()


col_widths = [22]*2 + [11]*(sku.shape[1] - 2)

wb = op.load_workbook(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))
side = op.styles.Side(border_style=None)
no_border = op.styles.borders.Border(left=side, right=side, top=side, bottom=side)
font_txt = op.styles.Font(name='Arial', size=10)
font_head = op.styles.Font(name='Arial', size=10, bold=True)
for ws in wb:
    for col in ws.iter_cols():
        ws.column_dimensions[op.utils.get_column_letter(col[0].column)].width = col_widths[col[0].column - 1]
        
        for cell in col:
            if cell.row == 1:
                cell.font = font_head
                cell.border = no_border
            else:
                cell.font = font_txt

            if cell.column > 2 and cell.row == 1:
                    cell.number_format = 'MMM YYYY'
            if cell.column > 2 and cell.row > 1:
                if ws.title == 'main':
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '0%'
wb.save(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))


print(f'{datetime.now().strftime("%H:%M:%S")}> Готово')