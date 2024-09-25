import numpy as np
import pandas as pd
np.set_printoptions(suppress=True, precision=2)
pd.options.display.float_format = '{:,.2f}'.format

import openpyxl as op
import duckdb

import os
import sys
from datetime import datetime


try:
    LOCAL_DIR = os.path.dirname(__file__)
except:
    LOCAL_DIR = os.getcwd()

PARENT_DIR = os.path.dirname(LOCAL_DIR)
sys.path.append(PARENT_DIR)

import tbl_calc

REPORT_MTH = tbl_calc.REPORT_MTH
DB_FILE = tbl_calc.DB_FILE


print('\n15. Бренды')
print('-'*79)


groups_dict = {
    'Группа 1': 'a. Группа 1',
    'Группа 2': 'b. Группа 2',
    'Группа 3': 'c. Группа 3',
    'Группа 4': 'd. Группа 4',
    'Группа 5': 'e. Группа 5',
    'Группа 6': 'f. Группа 6'
}

producers_dict = {
    'a. Группа 1': {
        'Производитель 1': 'a. Производитель 1',
        'Производитель 2': 'b. Производитель 2',
        'Производитель 3': 'c. Производитель 3',
        'Производитель 4': 'd. Производитель 4',
        'Производитель 5': 'e. Производитель 5',
        'Производитель 6': 'f. Производитель 6',
        'Производитель 7': 'g. Производитель 7',
        'Производитель 8': 'h. Производитель 8',
        'Прочие':          'i. Прочие'
    },

    'b. Группа 2': {
        'Производитель 1':  'a. Производитель 1',
        'Производитель 2':  'b. Производитель 2',
        'Производитель 3':  'c. Производитель 3',
        'Производитель 4':  'd. Производитель 4',
        'Производитель 5':  'e. Производитель 5',
        'Производитель 6':  'f. Производитель 6',
        'Производитель 7':  'g. Производитель 7',
        'Производитель 8':  'h. Производитель 8',
        'Производитель 9':  'i. Производитель 9',
        'Производитель 10': 'j. Производитель 10',
        'Прочие':           'k. Прочие'
    },

    'c. Группа 3': {
        'Производитель 1':  'a. Производитель 1',
        'Производитель 2':  'b. Производитель 2',
        'Производитель 3':  'c. Производитель 3',
        'Производитель 4':  'd. Производитель 4',
        'Производитель 5':  'e. Производитель 5',
        'Прочие':           'f. Прочие'
    },

    'd. Группа 4': {
        'Производитель 1':  'a. Производитель 1',
        'Производитель 2':  'b. Производитель 2',
        'Производитель 3':  'c. Производитель 3',
        'Производитель 4':  'd. Производитель 4',
        'Производитель 5':  'e. Производитель 5',
        'Производитель 6':  'f. Производитель 6',
        'Производитель 7':  'g. Производитель 7',
        'Производитель 8':  'h. Производитель 8',
        'Производитель 9':  'i. Производитель 9',
        'Производитель 10': 'j. Производитель 10',
        'Прочие':           'k. Прочие'
    },

    'e. Группа 5': {
        'Производитель 1':  'a. Производитель 1',
        'Производитель 2':  'b. Производитель 2',
        'Производитель 3':  'c. Производитель 3',
        'Производитель 4':  'd. Производитель 4',
        'Производитель 5':  'e. Производитель 5',
        'Производитель 6':  'f. Производитель 6',
        'Производитель 7':  'g. Производитель 7',
        'Производитель 8':  'h. Производитель 8',
        'Производитель 9':  'i. Производитель 9',
        'Производитель 10': 'j. Производитель 10',
        'Прочие':           'k. Прочие'
    },

     'f. Группа 6': {
        'Производитель 1':  'a. Производитель 1',
        'Производитель 2':  'b. Производитель 2',
        'Производитель 3':  'c. Производитель 3',
        'Производитель 4':  'd. Производитель 4',
        'Прочие':           'e. Прочие'
    }
}


try:
    conx = duckdb.connect(database=DB_FILE, read_only=True)
except Exception as err:
    print(f'Не удается подключиться к базе данных:\n{err}')


try:
    with open(os.path.join(LOCAL_DIR, '15. Бренды.sql'), encoding='utf-8') as q:
        sql_query = q.read()
except Exception as err:
    print(f'Не удается открыть файл с SQL запросом:\n{err}')


df = conx.execute(sql_query).fetch_df()
conx.close()


df.to_csv(os.path.join(LOCAL_DIR, 'data.csv'), index=False)


df['group'] = df['group'].replace(groups_dict)


grouped = df.groupby("group")


writer = pd.ExcelWriter(os.path.join(LOCAL_DIR, 'output.xlsx'))

for name, group in grouped:
    df_tmp = group.copy()

    # Замена наименований производителей и брендов для последующей сортировки по тоннажу и производителю
    df_tmp['producer'] = df_tmp['producer'].replace(producers_dict[name])
    df_tmp['brand'] = df_tmp['brand'].replace(producers_dict[name])

    # Создание сводной таблицы (строки: производитель, бренд; столбцы: месяцы)
    df_tn = df_tmp.pivot(index=['producer', 'brand'], columns='report_date', values='tonnes')
    df_rub = df_tmp.pivot(index=['producer', 'brand'], columns='report_date', values='thou_rub')

    # Сумма в тоннаже за последние полгода
    df_tn['total'] = df_tn.iloc[:, -6:].sum(axis=1)
    df_tn = df_tn.reset_index()
    
    # Добавление информации о тоннаже в датафрейм с рублями
    df_rub = df_rub.reset_index()
    df_rub = df_rub.merge(df_tn[['brand', 'total']], on='brand')

    # Сортировка двух таблиц по единому принципу 
    # (тоннаж за полгода - по убыванию, наименование производителя - по возрастанию)
    df_tn = df_tn.sort_values(axis='index', by=['producer', 'total'], ascending=[True, False])
    df_rub = df_rub.sort_values(axis='index', by=['producer', 'total'], ascending=[True, False])

    # Заполнение пропусков нулями
    df_tn = df_tn.fillna(0)
    df_rub = df_rub.fillna(0)

    # Удаление вспомагательного столбца с суммой кг за полгода
    df_tn = df_tn.drop(columns='total')
    df_rub = df_rub.drop(columns='total')

    # Очистка наименований производителей и брендов от вспомагательной информации
    df_tn[['producer', 'brand']] = df_tn[['producer', 'brand']].replace(r'^\w\S\s', '', regex=True)
    df_rub[['producer', 'brand']] = df_rub[['producer', 'brand']].replace(r'^\w\S\s', '', regex=True)

    # Установка столбцов в качестве индексов
    df_tn = df_tn.set_index(['producer', 'brand'])
    df_rub = df_rub.set_index(['producer', 'brand'])

    # Копии датафреймов для расчета долей
    df_tn_pct = df_tn.copy()
    df_rub_pct = df_rub.copy()
    
    # Делить помесячные данные на 1/2 суммы по производителям и брендам плюс сумму по XXX и прочим производителям
    df_tn_pct = df_tn.div(df_tn.iloc[0:-2, :].sum()/2 + df_tn.iloc[-2:, :].sum(), axis=1, fill_value=None)*100
    df_rub_pct = df_rub.div(df_rub.iloc[0:-2, :].sum()/2 + df_rub.iloc[-2:, :].sum(), axis=1, fill_value=None)*100 
    
    # Средние цены
    df_price = df_rub / df_tn

    # Итоги для абсолютных показателей
    df_tn = pd.concat([
                df_tn,
                pd.DataFrame(
                    data=(df_tn.iloc[0:-2, :].sum()/2 + df_tn.iloc[-2:, :].sum()).values.reshape(1, -1), 
                    index=pd.MultiIndex.from_tuples([('Итого', 'Итого')], names=['producer', 'brand']), 
                    columns=df_tn.columns)
            ])
    df_rub = pd.concat([
                df_rub,
                pd.DataFrame(
                    data=(df_rub.iloc[0:-2, :].sum()/2 + df_rub.iloc[-2:, :].sum()).values.reshape(1, -1), 
                    index=pd.MultiIndex.from_tuples([('Итого', 'Итого')], names=['producer', 'brand']), 
                    columns=df_rub.columns)
            ])

    # Расчет отклонений
    df_tn = (pd.concat([
            df_tn,
            df_tn.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m',     calc_method='pct_change', aggfunc='sum').rename('m_on_m_pct'),
            df_tn.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='ytd_on_ytd', calc_method='pct_change', aggfunc='sum').rename('ytd_on_ytd_pct')
        ] , axis=1)
        .reset_index(names=['producer', 'brand']).replace(r'^\w\S\s', '', regex=True)
        .pipe(tbl_calc.insert_wsp, 2)
    )
    df_rub = (pd.concat([
            df_rub,
            df_rub.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m',     calc_method='pct_change', aggfunc='sum').rename('m_on_m_pct'),
            df_rub.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='ytd_on_ytd', calc_method='pct_change', aggfunc='sum').rename('ytd_on_ytd_pct')
        ] , axis=1)
        .reset_index(names=['producer', 'brand']).replace(r'^\w\S\s', '', regex=True)
        .pipe(tbl_calc.insert_wsp, 2)
    )

    df_tn_pct = (pd.concat([
            df_tn_pct,
            df_tn_pct.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m',     calc_method='pp_change', aggfunc='mean').rename('m_on_m_abs'),
            df_tn_pct.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='ytd_on_ytd', calc_method='pp_change', aggfunc='mean').rename('ytd_on_ytd_abs')
        ] , axis=1)
        .reset_index(names=['producer', 'brand']).replace(r'^\w\S\s', '', regex=True)
        .pipe(tbl_calc.insert_wsp, 2)
    )

    df_rub_pct = (pd.concat([
            df_rub_pct,
            df_rub_pct.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m',     calc_method='pp_change', aggfunc='mean').rename('m_on_m_abs'),
            df_rub_pct.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='ytd_on_ytd', calc_method='pp_change', aggfunc='mean').rename('ytd_on_ytd_abs')
        ] , axis=1)
        .reset_index(names=['producer', 'brand']).replace(r'^\w\S\s', '', regex=True)
        .pipe(tbl_calc.insert_wsp, 2)
    )

    df_price = (pd.concat([
            df_price,
            df_price.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m',     calc_method='pct_change', aggfunc='mean').rename('m_on_m_pct'),
            df_price.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='ytd_on_ytd', calc_method='pct_change', aggfunc='mean').rename('ytd_on_ytd_pct')
        ] , axis=1)
        .reset_index(names=['producer', 'brand']).replace(r'^\w\S\s', '', regex=True).fillna('–')
        .pipe(tbl_calc.insert_wsp, 2)
    )

    # Сохранение данных
    df_tn_pct.to_excel(writer, sheet_name=f'{name[3:]}_тн_%', index=False)
    df_tn.to_excel(writer, sheet_name=f'{name[3:]}_тн', index=False)
    df_rub_pct.to_excel(writer, sheet_name=f'{name[3:]}_руб_%', index=False)
    df_rub.to_excel(writer, sheet_name=f'{name[3:]}_руб', index=False)
    df_price.to_excel(writer, sheet_name=f'{name[3:]}_цена', index=False)

writer.close()


# Базовое форматирование
col_widths = [25]*2 + [12]*(df_tn.shape[1] - 2)

wb = op.load_workbook(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))
side = op.styles.Side(border_style=None)
no_border = op.styles.borders.Border(left=side, right=side, top=side, bottom=side)
font_txt = op.styles.Font(name='Arial', size=10)
font_head = op.styles.Font(name='Arial', size=10, bold=True)
for ws in wb:
    for col in ws.iter_cols():
        ws.column_dimensions[op.utils.get_column_letter(col[0].column)].width = col_widths[col[0].column - 1]
        for cell in col:
            if 2 < cell.column < (df_tn.shape[1] - 2) and cell.row == 1:
                cell.number_format = 'MMM YYYY'
            if 2 < cell.column < (df_tn.shape[1] - 2) and cell.row > 1:
                    if ('тн_%' in ws.title or 'руб_%' in ws.title or '_цена' in ws.title):
                        cell.number_format = '#,##0.0'
                    else:
                        cell.number_format = '#,##0'
            if cell.row == 1:
                cell.font = font_head
                cell.border = no_border
            else:
                cell.font = font_txt
wb.save(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))


print(f'{datetime.now().strftime("%H:%M:%S")}> Готово')