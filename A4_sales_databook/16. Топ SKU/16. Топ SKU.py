import numpy as np
import pandas as pd
np.set_printoptions(suppress=True, precision=2)
pd.options.display.float_format = '{:,.2f}'.format

import openpyxl as op
import duckdb

import os
import sys
from datetime import datetime


try:
    LOCAL_DIR = os.path.dirname(__file__)
except:
    LOCAL_DIR = os.getcwd()

PARENT_DIR = os.path.dirname(LOCAL_DIR)
sys.path.append(PARENT_DIR)

import tbl_calc

REPORT_MTH = tbl_calc.REPORT_MTH
DB_FILE = tbl_calc.DB_FILE


def pretty_change(row):
    try:
        if row["tn_curr"]/row["tn_prev"]-1 > 1.5:
            chn_tn = f'{row["tn_curr"]/row["tn_prev"]:.1f}x'
        else:
            chn_tn = f'{row["tn_curr"]/row["tn_prev"]-1:.1%}'
    except:
        chn_tn = '–'

    try:
        if row["thou_rub_curr"]/row["thou_rub_prev"]-1 > 1.5:
            chn_rub = f'{row["thou_rub_curr"]/row["thou_rub_prev"]:.1f}x'
        else:
            chn_rub = f'{row["thou_rub_curr"]/row["thou_rub_prev"]-1:.1%}'
    except:
        chn_rub = '–'

    return pd.Series({
        'chn_tn': chn_tn.replace('.', ','),
        'chn_rub': chn_rub.replace('.', ','),
    })


print('\n16. Топ SKU')
print('-'*79)


cols_final = ['num', 'global_sku_name', 'brand', 'weight',
              'tn_prev', 'tn_curr', 'chn_tn', 'sep1',
              'thou_rub_prev', 'thou_rub_curr', 'chn_rub', 'sep2',
              'avr_price_prev', 'avr_price_curr', 'sep3',
              'akb_prev', 'akb_curr', 'sep4',
              'turnover_kg_mnth_prev', 'turnover_kg_mnth_curr'
]


groups_dict = {
    'Группа 1': 'a. Группа 1',
    'Группа 2': 'b. Группа 2',
    'Группа 3': 'c. Группа 3',
    'Группа 4': 'd. Группа 4',
    'Группа 5': 'e. Группа 5',
    'Группа 6': 'f. Группа 6'
}

store_formats_dict = {
    'все': ['Формат 1', 'Формат 2', 'Формат 3', 'Формат 4', 'Прочие'],
    'X1': ['Формат 1'],
    'X2': ['Формат 2'],
    'X3': ['Формат 3'],
    'X4': ['Формат 4']
}


try:
    conx = duckdb.connect(database=DB_FILE, read_only=True)
except Exception as err:
    print(f'Не удается подключиться к базе данных:\n{err}')


try:
    with open(os.path.join(LOCAL_DIR, '16. Топ SKU.sql'), encoding='utf-8') as q:
        sql_query = q.read()
except Exception as err:
    print(f'Не удается открыть файл с SQL запросом:\n{err}')


for k, v in store_formats_dict.items():
    writer = pd.ExcelWriter(os.path.join(LOCAL_DIR, f'output ({k}).xlsx'))

    df = conx.execute(sql_query, [v, v]).fetch_df()
    df['group'] = df['group'].replace(groups_dict)

    grouped = df.groupby('group')

    for name, group in grouped:
        df_tmp = group.copy()
        df_tmp = df_tmp.drop(columns='group')

        first_row = pd.DataFrame({
                'num': '',
                'global_sku_name': 'Все SKU',
                'producer': '',
                'brand': '',
                'weight': '',
                'tn_prev': df_tmp['tn_prev'].sum(),
                'tn_curr': df_tmp['tn_curr'].sum(),
                'thou_rub_prev': df_tmp['thou_rub_prev'].sum(),
                'thou_rub_curr': df_tmp['thou_rub_curr'].sum(),
                'avr_price_prev': df_tmp['thou_rub_prev'].sum() / df_tmp['tn_prev'].sum(),
                'avr_price_curr': df_tmp['thou_rub_curr'].sum() / df_tmp['tn_curr'].sum(),
                'akb_prev': '–',
                'akb_curr': '–',
                'turnover_kg_mnth_prev': '–',
                'turnover_kg_mnth_curr': '–'
            },
            index=[0]
        )

        if df_tmp.shape[0] > 50:
            others_row = pd.DataFrame({
                    'num': '',
                    'global_sku_name': 'Прочие, в том числе:',
                    'producer': '',
                    'brand': '',
                    'weight': '',
                    'tn_prev': df_tmp.iloc[50:, ].loc[:, 'tn_prev'].sum(),
                    'tn_curr': df_tmp.iloc[50:, ].loc[:, 'tn_curr'].sum(),
                    'thou_rub_prev': df_tmp.iloc[50:, ].loc[:, 'thou_rub_prev'].sum(),
                    'thou_rub_curr': df_tmp.iloc[50:, ].loc[:, 'thou_rub_curr'].sum(),
                    'avr_price_prev': df_tmp.iloc[50:, ].loc[:, 'thou_rub_prev'].sum() / df_tmp.iloc[50:, ].loc[:, 'tn_prev'].sum(),
                    'avr_price_curr': df_tmp.iloc[50:, ].loc[:, 'thou_rub_curr'].sum() / df_tmp.iloc[50:, ].loc[:, 'tn_curr'].sum(),
                    'akb_prev': '–',
                    'akb_curr': '–',
                    'turnover_kg_mnth_prev': '–',
                    'turnover_kg_mnth_curr': '–'
                },
                index=[0]
            )

            others_xx = df_tmp.iloc[50:200].loc[df_tmp.iloc[50:200].loc[:, 'producer'] == 'XXXX']
            others_xx = others_xx.iloc[0:5, :]
            df_tmp = pd.concat([first_row, df_tmp[0:50], others_row, others_xx])
        else:
            df_tmp = pd.concat([first_row, df_tmp[0:50]])

        df_tmp = df_tmp.drop(columns='producer')

        df_tmp[['chn_tn', 'chn_rub']] = df_tmp.apply(pretty_change, axis=1)
        df_tmp[['sep1', 'sep2', 'sep3', 'sep4']] = None
        df_tmp = df_tmp.reindex(columns=cols_final)

        df_tmp.to_excel(writer, sheet_name=name[3:], index=False)

    writer.close()

    col_widths = [10, 68, 15, 6] + [12]*3 + [2] + [12]*3 + [2] + [12]*2 + [2] + [12]*2 + [2] + [12]*2

    wb = op.load_workbook(filename=os.path.join(LOCAL_DIR, f'output ({k}).xlsx'))
    side = op.styles.Side(border_style=None)
    no_border = op.styles.borders.Border(left=side, right=side, top=side, bottom=side)
    font_txt = op.styles.Font(name='Arial', size=10)
    font_head = op.styles.Font(name='Arial', size=10, bold=True)
    for ws in wb:
        for col in ws.iter_cols():
            ws.column_dimensions[op.utils.get_column_letter(col[0].column)].width = col_widths[col[0].column - 1]
            for cell in col:
                if cell.row == 1:
                    cell.font = font_head
                    cell.border = no_border
                else:
                    cell.font = font_txt

                if cell.column == 1:
                    cell.alignment = op.styles.Alignment(horizontal='center')
                elif cell.column == 4:
                    cell.alignment = op.styles.Alignment(horizontal='left')
                elif cell.column > 4:
                    cell.alignment = op.styles.Alignment(horizontal='right')

                if cell.column in [5, 6, 9, 10, 16, 17]:
                    cell.number_format = '#,##0'
                elif cell.column in [13, 14, 19, 20]:
                    cell.number_format = '#,##0.0'

    wb.save(filename=os.path.join(LOCAL_DIR, f'output ({k}).xlsx'))


conx.close()


print(f'{datetime.now().strftime("%H:%M:%S")}> Готово')