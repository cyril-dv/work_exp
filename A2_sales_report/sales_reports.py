import numpy as np
import pandas as pd

import psycopg2
import openpyxl as op

import sys
import locale
import os
from dateutil import relativedelta
from datetime import datetime, date
from itertools import zip_longest

np.set_printoptions(suppress=True, precision=4)
pd.options.display.float_format = '{:,.4f}'.format
idx = pd.IndexSlice

locale.setlocale(locale.LC_ALL, 'ru_RU.utf8');


def wk_iso_num(dt):
    if not isinstance(dt, datetime):
        raise Exception(f'Expected datetime, got "{type(dt).__name__}" instead')
    else:
        yr = str(dt.isocalendar()[0])
        wk = str(dt.isocalendar()[1])
        if len(wk) == 1:
            wk = '0' + wk
        return yr + '-' + wk


# Текущая директория и директория, в которой будет храниться результирующий файл
local_dir = os.getcwd()
res_subdir = 'output'

# Префикс с текущей датой
curr_date = str(date.today())

# Индикатор основного цикла
is_running = True


print()
print('-'*75)
print('Отчеты sales report')
print('-'*75)


while is_running:
    # Детали для подключения к PostgreSQL
    try:
        conx = psycopg2.connect(
                host='...',
                dbname='...',
                user='...',
                password='...'
        )
    except (Exception, psycopg2.DatabaseError) as err:
        raise Exception('Не удается подключиться к базе данных!') from err


    # Запрос для выгрузки информации из представления sales_report
    main_query = "select * from sales_report where network_subname = %s"

    # Запрос на получение списка сетей/форматов
    nets_query = "select distinct network_subname from sales_report order by 1"


    # Список сетей, для которых присутствуют агрегированные данные в таблице
    curs = conx.cursor()
    curs.execute(nets_query)
    networks = [i[0] for i in curs.fetchall()]
    networks_dict = {k:v for k, v in enumerate(networks, start=1)}
    networks_split = np.int32(np.ceil(len(networks)/2))

    print()
    print('Доступные сети:')
    print(
        "".join(
            f"{net[0] + ' [' + str(num) + ']':35}{net[1]} [{num+networks_split if num+networks_split <= len(networks) else '--'}]\n"
            for num, net in enumerate(zip_longest(networks[0:networks_split], networks[networks_split:], fillvalue="-----"), start=1)
            )
        )


    # Выбор пользователем сети для подготовки отчета
    while True:
        try:
            chosen_net = int(input('Выберите номер сети или 0 для завершения: '))
        except:
            print('Указан неверный номер сети.')
            continue

        if chosen_net in networks_dict.keys():
            break
        elif chosen_net == 0:
            curs.close()
            conx.close()
            sys.exit()
        else:
            print('Указан неверный номер сети.')

    chosen_net = networks_dict[chosen_net]


    # Особый кейс для понедельных данных для ____ в разрезе корпоративных центров
    if 'понед.' in chosen_net:
        data_frq = 'Неделя'
    else:
        data_frq = 'Месяц'


    # Получение данных для выбраной сети
    curs.execute(main_query, (chosen_net,))
    df = pd.DataFrame(curs.fetchall())

    try:
        df.columns = [desc[0] for desc in curs.description]
    except Exception as err:
        raise Exception('Запрос не вернул данных!') from err


    print(f'Получено {df.shape[0]:,d} строк из базы данных.')


    # Имитирование null значений в DBeaver
    df['report_date'] = pd.to_datetime(df['report_date'])
    df = df.where(pd.notnull(df), '[NULL]')


    # Названия строк и столбцов итоговой сводной таблицы
    cols_idx = ['Категория', 'Сегмент', 'Товарная группа', 'Бренд', 'Производитель', 'Наименование', 'Вес']
    cols_mnts = [data_frq]
    cols_vals = ['АКБ', 'Оборачиваемость, шт/ТТ', 'РТО, руб/ТТ', 'ТО, шт', 'ТО, кг', 'Цена за шт', 'Цена за кг', 'ТО, руб.']


    # Информация для заголовка сводной таблицы
    chain = f"Динамика основных показателей по продажам в {df.at[0, 'network_subname']}"
    dates = f"{np.sort(pd.DatetimeIndex(df['report_date'].unique()).to_pydatetime())[0].strftime('%B %Y')} - {np.sort(pd.DatetimeIndex(df['report_date'].unique()).to_pydatetime())[-1].strftime('%B %Y')}"
    df = df.drop(columns='network_subname')


    # Переименование столбцов
    df = df.rename(
        columns={
            'report_date': data_frq, 
            'category': 'Категория', 
            'segment': 'Сегмент', 
            'group': 'Товарная группа', 
            'brand': 'Бренд',
            'producer': 'Производитель', 
            'global_sku_name': 'Наименование', 
            'weight': 'Вес', 
            'akb': 'АКБ',
            'pieces': 'ТО, шт',
            'kg': 'ТО, кг',
            'sales_rub': 'ТО, руб.'
        }
    )


    # Сортировка исходных данных в хронологическом порядке (для корректного отражения второго уровня мультииндекса столбцов)
    df = df.sort_values(by=data_frq)


    # Расчет производных показателей
    df['Оборачиваемость, шт/ТТ'] = df['ТО, шт'] / df['АКБ']
    df['РТО, руб/ТТ'] = df['ТО, руб.'] / df['АКБ']
    df['Цена за шт'] = df['ТО, руб.'] / df['ТО, шт']
    df['Цена за кг'] = df['ТО, руб.'] / df['ТО, кг']


    # Для SKU, у которых не указан вес, установить цену за кг, равную нулю
    df['Цена за кг'] = df['Цена за кг'].replace({np.inf:0})


    # Вывод показателей помесячно в столбцы
    df = df.pivot_table(
        index=cols_idx,
        columns=cols_mnts, 
        values=cols_vals,
        aggfunc='max',
        sort=False)


    # Требуемый порядок столбцов с показателями (иначе в алфавитном порядке)
    df = df.reindex(columns=cols_vals, level=0)


    # Округление значений всех показателей
    df = df.round(2)


    # Определение месяцев для расчета продаж за последний месяц и последние три месяца
    month_last = df.loc[:, idx['АКБ']].columns[-1].date()

    if data_frq == 'Неделя':
        month_start_3 = month_last + relativedelta.relativedelta(weeks=-12)
    else:
        month_start_3 = month_last + relativedelta.relativedelta(months=-2)

    month_start_all = df.loc[:, idx['АКБ']].columns[0].date()

    month_last = pd.Timestamp(month_last)
    month_start_3 = pd.Timestamp(month_start_3)
    month_start_all = pd.Timestamp(month_start_all)


    # Столбец с данными продаж в руб. за последние три месяца и за все время наблюдений для сортировки
    df[('Сортировка 3м', 'по руб.')] = (
        df.loc[:, idx['ТО, руб.', month_start_3:month_last]]
        .loc[:, 'ТО, руб.']
        .sum(axis=1)
    )

    df[('Сортировка все_м', 'по руб.')] = (
        df.loc[:, idx['ТО, руб.', month_start_all:month_last]]
        .loc[:, 'ТО, руб.']
        .sum(axis=1)
    )


    # Установить имя для индекса столбцов
    df.columns.set_names(names=['Индекс', 'Показатель'], inplace=True)


    # Перемещение значений уровней индекса строк в данные датафрейма, установка названий для второго уровня индекса столбцов
    df = df.reset_index(col_fill='Показатель')


    # Маски для поиска строк с группами и подгруппами
    # lv_1 - общий итог
    # lv_2 - подытог по товарной группе
    # lv_3 - подытог по производителю
    mask_lv1 = (df[('Категория', 'Показатель')] == '[NULL]')
    mask_lv2 = ((df[('Бренд', 'Показатель')] == '[NULL]') & (df[('Производитель', 'Показатель')] == '[NULL]') & (df[('Товарная группа', 'Показатель')] != '[NULL]'))
    mask_lv3 = ((df[('Бренд', 'Показатель')] == '[NULL]') & (df[('Наименование', 'Показатель')] == '[NULL]') & (df[('Производитель', 'Показатель')] != '[NULL]'))


    # Проставление в столбце "Наименование" информации об итоге, товарной группе и прозводителе
    df.loc[mask_lv1, ('Наименование', 'Показатель')] = 'Общий итог'
    df.loc[mask_lv2, ('Наименование', 'Показатель')] = df[('Товарная группа', 'Показатель')]
    df.loc[mask_lv3, ('Наименование', 'Показатель')] = df[('Производитель', 'Показатель')]


    # Сортировка по общему итогу (0), товарной группе (1), производителю (2) и SKU (3)
    df[('Сортировка пр-тель', 'по группе')] = np.where(mask_lv2, 1, np.where(mask_lv3, 2, 3))
    df.loc[(df[('Категория', 'Показатель')] == '[NULL]'), ('Сортировка пр-тель', 'по группе')] = 0


    # Сортировка данных внутри датафрейма
    df = df.sort_values(
        axis='index',
        by=[('Категория', 'Показатель'), ('Сегмент', 'Показатель'), ('Товарная группа', 'Показатель'), ('Сортировка пр-тель', 'по группе'), ('Сортировка 3м', 'по руб.'), ('Сортировка все_м', 'по руб.'), ('Наименование', 'Показатель')], 
        ascending=[True, True, True, True, False, False, True])


    # Временная таблица с данными по топ производителям
    df_producers = (df.loc[
                        df[('Сортировка пр-тель', 'по группе')] == 2,
                        [('Товарная группа', 'Показатель'), ('Производитель', 'Показатель'), ('Сортировка 3м', 'по руб.')]
                    ]
                    .droplevel(1, axis='columns')
                    .sort_values(by='Сортировка 3м')
    )
    grouped = df_producers.groupby('Товарная группа')['Сортировка 3м']
    df_producers['cumsum'] = grouped.transform(pd.Series.cumsum)

    grouped = df_producers.groupby('Товарная группа')['cumsum']
    df_producers['top_producers'] = grouped.transform(lambda x: (x >= 0.01 * x.iat[-1]))

    misc_producers = df_producers.loc[df_producers['top_producers'] == False, :].index


    # Удаление строк с итогами по производителям, которые не входят в топ перечень
    df = df.drop(index=misc_producers)

    # Создание массива с индексами строк в финальном excel файле (группировка строк) для SKU, 
    # продажи которых составляют менее 1% от продаж в категории за последние три месяца. 
    # Для продукции ____ действует правило: продажи составляеют менее ____ тыс. рублей за последние три месяца.
    grouped = df.groupby([('Категория', 'Показатель'), ('Сегмент', 'Показатель'), ('Товарная группа', 'Показатель')])
    hide_idx = np.array([])

    for name, group in grouped:
        df_tmp = group.reset_index()
        group_total = df_tmp[('Сортировка 3м', 'по руб.')].iat[0]
        grp_idx = np.zeros(df_tmp.shape[0])

        for i, r in df_tmp.iterrows():
            if (((r[('Сортировка пр-тель', 'по группе')] == 3) and (r[('Производитель', 'Показатель')] != '____') and (r[('Сортировка 3м', 'по руб.')] < 0.01*group_total)) 
                or ((r[('Сортировка пр-тель', 'по группе')] == 3) and (r[('Производитель', 'Показатель')] == '____') and (r[('Сортировка 3м', 'по руб.')] < 100))):
                grp_idx[i] = 1

        hide_idx = np.r_[hide_idx, grp_idx]

    # Индексирование строк в openpyxl начинается с одного; отступ в пять строк перед таблицей в excel файле
    hide_idx = np.nonzero(hide_idx)[0] + 1 + 5


    # Удаление столбцов с информацией для сортировки
    df = df.drop(columns=[('Сортировка пр-тель', 'по группе'), ('Сортировка 3м', 'по руб.'), ('Сортировка все_м', 'по руб.')])


    # Удаление текста с NULL значениями
    df = df.replace('[NULL]', '')


    # Создать директорию output, если она не существует
    if not os.path.isdir(os.path.join(local_dir, res_subdir)):
        os.mkdir(os.path.join(local_dir, res_subdir))


    # Сохранить сводную таблицу без форматирования
    df.to_excel(os.path.join(local_dir, res_subdir, f'{chosen_net}_{curr_date}.xlsx'), merge_cells=True)


    print(f'Файл "{chosen_net}_{curr_date}.xlsx" сохранен в папке output.')


    # Количество строк и столбцов в сводной таблице
    row_total = df.shape[0] + 2                         # Общее кол-во строк
    col_total = df.shape[1] + 1                         # Общее кол-во столбцов
    col_meta = 8                                        # Кол-во столбцов с информацией о товаре
    col_meta_width = [8, 16, 27, 42, 12, 30, 40, 8]     # Ширина столбцов с информацией о товаре
    col_data = col_total - col_meta                     # Кол-во столбцов с числовыми данными
    col_periods = int(col_data / len(cols_vals))        # Кол-во отчетных периодов


    # Секционирование столбцов по показателям 
    # (номер начального и конечного столбца для каждого из показателей)
    col_sections = np.array([[i, i + col_periods - 1] for i in range(col_meta + 1, col_total + 1, col_periods)])


    wb = op.load_workbook(filename=os.path.join(local_dir, 'output', f'{chosen_net}_{curr_date}.xlsx'))
    ws = wb.active


    # Удалить пустую строку после заголовков
    ws.delete_rows(df.columns.nlevels + 1)


    # Форматирование заголовков описательных столбцов
    for col in ws.iter_cols(min_row=1, max_row=2, min_col=1, max_col=col_meta):
        # Ширина
        ws.column_dimensions[op.utils.get_column_letter(col[0].column)].width = col_meta_width[col[0].column - 1]
        # Объединение
        ws.merge_cells(start_column=col[0].column, end_column=col[0].column, start_row=1, end_row=2)
        # Форматирование
        ws.cell(1, col[0].column).alignment = op.styles.Alignment(horizontal='center', vertical='center')
        # Заливка
        ws.cell(1, col[0].column).fill = op.styles.PatternFill("solid", start_color="92D050")
        ws.cell(2, col[0].column).fill = op.styles.PatternFill("solid", start_color="92D050")

    ws.cell(1, 1).value = '№ п/п'

    # Очистка первого столбца
    for row in ws.iter_rows(min_row=3, max_row=row_total, min_col=1, max_col=1):
        row[0].value = None
        row[0].font = op.styles.Font(bold=False)


    # Перенос строк для столбцов "Производитель" и "Наименование"
    for row in ws.iter_rows(min_row=3, max_row=row_total, min_col=6, max_col=7):
        for cell in row:
            cell.alignment = op.styles.Alignment(wrap_text=True)


    # Форматирование заголовков численных столбцов
    for col in ws.iter_cols(min_row=1, max_row=2, min_col=col_meta+1, max_col=col_total):
        for cell in col:
            if cell.row == 1:
                ws.cell(1, cell.column).fill = op.styles.PatternFill("solid", start_color="FFC000")
            elif cell.row == 2:
                ws.cell(2, cell.column).fill = op.styles.PatternFill("solid", start_color="92D050")
                if data_frq == 'Неделя':
                    ws.cell(2, cell.column).value = wk_iso_num(ws.cell(2, cell.column).value)
                    ws.cell(2, cell.column).number_format = 'General'
                else:
                    ws.cell(2, cell.column).number_format = 'MMM YY'     


    # Форматирование численных столбцов
    sm_width = 11
    lg_width = 13

    for col in ws.iter_cols(min_row=3, max_row=row_total, min_col=col_meta+1, max_col=col_total):
        for cell in col:
            if (col_sections[0][0] <= cell.column <= col_sections[0][1]):
                ws.column_dimensions[op.utils.get_column_letter(cell.column)].width = sm_width
                ws.cell(cell.row, cell.column).number_format = '#,##0'
            elif col_sections[1][0] <= cell.column <= col_sections[1][1]:
                ws.column_dimensions[op.utils.get_column_letter(cell.column)].width = sm_width
                ws.cell(cell.row, cell.column).number_format = '#,##0.0'
            elif (col_sections[2][0] <= cell.column <= col_sections[2][1]):
                ws.column_dimensions[op.utils.get_column_letter(cell.column)].width = sm_width
                ws.cell(cell.row, cell.column).number_format = '#,##0'
            elif (col_sections[3][0] <= cell.column <= col_sections[3][1]):
                ws.column_dimensions[op.utils.get_column_letter(cell.column)].width = sm_width
                ws.cell(cell.row, cell.column).number_format = '#,##0'
            elif (col_sections[4][0] <= cell.column <= col_sections[4][1]):
                ws.column_dimensions[op.utils.get_column_letter(cell.column)].width = sm_width
                ws.cell(cell.row, cell.column).number_format = '#,##0'
            elif col_sections[5][0] <= cell.column <= col_sections[5][1]:
                ws.column_dimensions[op.utils.get_column_letter(cell.column)].width = sm_width
                ws.cell(cell.row, cell.column).number_format = '#,##0.0'
            elif col_sections[6][0] <= cell.column <= col_sections[6][1]:
                ws.column_dimensions[op.utils.get_column_letter(cell.column)].width = sm_width
                ws.cell(cell.row, cell.column).number_format = '#,##0.0'
            elif col_sections[7][0] <= cell.column <= col_sections[7][1]:
                ws.column_dimensions[op.utils.get_column_letter(cell.column)].width = lg_width
                ws.cell(cell.row, cell.column).number_format = '#,##0'
            elif col_sections[8][0] <= cell.column <= col_sections[8][1]:
                ws.column_dimensions[op.utils.get_column_letter(cell.column)].width = sm_width
                ws.cell(cell.row, cell.column).number_format = '#,##0.0'
            elif col_sections[9][0] <= cell.column <= col_sections[9][1]:
                ws.column_dimensions[op.utils.get_column_letter(cell.column)].width = sm_width
                ws.cell(cell.row, cell.column).number_format = '#,##0.0'
            elif col_sections[10][0] <= cell.column <= col_sections[10][1]:
                ws.column_dimensions[op.utils.get_column_letter(cell.column)].width = lg_width
                ws.cell(cell.row, cell.column).number_format = '#,##0'


    # Границы для всей таблицы
    thin_border  = op.styles.Border(
        left=op.styles.Side(style='thin'), 
        right=op.styles.Side(style='thin'), 
        top=op.styles.Side(style='thin'), 
        bottom=op.styles.Side(style='thin')
    )
    thick_border = op.styles.Border(
        left=op.styles.Side(style='thin'), 
        right=op.styles.Side(style='medium'), 
        top=op.styles.Side(style='thin'), 
        bottom=op.styles.Side(style='thin')
    )

    for col in ws.iter_cols(min_row=1, max_row=row_total, min_col=1, max_col=col_total):
        for cell in col:
            if cell.column in col_sections[:, 1]:
                cell.border = thick_border
            else:
                cell.border = thin_border


    # Заливка строк, определение позиций строк с заголовками
    subheaders_lv2 = set()
    subheaders_lv3 = set()

    for row in ws.iter_rows(min_row=3, max_row=row_total, min_col=2, max_col=col_total):
        for cell in row:
            if cell.column == 4 and cell.value is None:
                for cell2 in ws[cell.row]:
                    ws.cell(cell2.row, cell2.column).fill = op.styles.PatternFill("solid", start_color="FFFF00")
                    ws.cell(cell2.row, cell2.column).font = op.styles.Font(size=12, bold=True)
            elif cell.column == 6 and cell.value is None and cell.offset(column=-2).value is not None:
                for cell2 in ws[cell.row]:
                    subheaders_lv2.add(cell.row)
                    ws.cell(cell2.row, cell2.column).fill = op.styles.PatternFill("solid", start_color="FFC000")
                    ws.cell(cell2.row, cell2.column).font = op.styles.Font(size=12, bold=True)
            elif cell.column == 5 and cell.value is None and cell.offset(column=-1).value is not None:
                for cell2 in ws[cell.row]:
                    subheaders_lv3.add(cell.row)
                    ws.cell(cell2.row, cell2.column).fill = op.styles.PatternFill("solid", start_color="FCE4D6")
                    ws.cell(cell2.row, cell2.column).font = op.styles.Font(size=11, bold=True)


    # Нумерация строк с SKU
    subheaders_lv2 = np.sort(np.array(list(subheaders_lv2)))
    subheaders_lv3 = np.sort(np.array(list(subheaders_lv3)))
    rowstonum = np.array([
        [np.append(subheaders_lv2, row_total)[i], np.append(subheaders_lv2, row_total)[i+1]] for i in range(len(np.append(subheaders_lv2, row_total))-1)
        ])

    for i in rowstonum:
        n = 1
        for j in range(i[0], i[1]+1):
            if j not in subheaders_lv2 and j not in subheaders_lv3:
                ws.cell(j, 1).value = n
                n += 1


    # Нумерация строк с производителями
    roman_nums = [
        'I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII',	'IX', 'X',
        'XI', 'XII', 'XIII', 'XIV', 'XV', 'XVI', 'XVII', 'XVIII', 'XIX', 'XX',
        'XXI', 'XXII', 'XXIII', 'XXIV', 'XXV', 'XXVI', 'XXVII', 'XXVIII', 'XXIX', 'XXX',
        'XXXI', 'XXXII', 'XXXIII', 'XXXIV', 'XXXV', 'XXXVI', 'XXXVII', 'XXXVIII', 'XXXIX', 'XL', 
        'XLI', 'XLII', 'XLIII', 'XLIV', 'XLV', 'XLVI', 'XLVII', 'XLVIII', 'XLVIX', 'L'
    ]

    for i in rowstonum:
        n = 0
        for j in range(i[0], i[1]):
            if j in subheaders_lv3 and j not in subheaders_lv2:
                ws.cell(j, 1).value = roman_nums[n]
                n += 1


    # Выделение строк с SKU от ____
    for col in ws.iter_cols(min_row=3, max_row=row_total, min_col=1, max_col=col_total):
        for cell in col:
            if ws.cell(cell.row, 1).value is not None and ws.cell(cell.row, 5).value is not None and ws.cell(cell.row, 6).value == '____':
                ws.cell(cell.row, cell.column).fill = op.styles.PatternFill("solid", start_color="FFFF99")


    # Добавить строки над основной таблицей
    ws.insert_rows(idx=0, amount=3)

    # Исправить образовавшиеся объединенные ячейки сверху таблицы
    for i in range(col_meta):
        ws.merge_cells(start_row=1, start_column=i+1, end_row=2, end_column=i+1)
        ws.unmerge_cells(start_row=1, start_column=i+1, end_row=2, end_column=i+1)

    for i in range(len(cols_vals)):
        ws.merge_cells(start_row=1, start_column=col_sections[i][0], end_row=1, end_column=col_sections[i][1])
        ws.unmerge_cells(start_row=1, start_column=col_sections[i][0], end_row=1, end_column=col_sections[i][1])

    # Исправить исчезвнувшее объединение ячеек в заголовках таблицы
    for i in range(col_meta):
        ws.merge_cells(start_row=4, start_column=i+1, end_row=5, end_column=i+1)

    for i in range(len(cols_vals)):
        ws.merge_cells(start_row=4, start_column=col_sections[i][0], end_row=4, end_column=col_sections[i][1])


    # Заголовок таблицы
    ws.cell(1, 7).value = chain
    ws.cell(1, 7).font = op.styles.Font(bold=True)
    ws.cell(2, 7).value = dates


    # Скрытие строк с SKU, у которых объем продаж составляет менее 1% от группы за последние 3 месяца
    for i in hide_idx:
        ws.row_dimensions.group(int(i), int(i), outline_level=1)


    wb.save(filename=os.path.join(local_dir, res_subdir, f'{chosen_net}_{curr_date}.xlsx'))


    print(f'В файл добавлено форматирование.')
    print()


    while True:
        is_running_chk = input('Создать еще один отчет? [Y/N] ').upper()
        if is_running_chk in ['Y', 'N']:
            break
        else:
            print('Получен некорректный ответ.')

    if is_running_chk == 'Y':
        is_running = True
    else:
        is_running = False

curs.close()
conx.close()

print()
print('Работа скрипта завершена.')