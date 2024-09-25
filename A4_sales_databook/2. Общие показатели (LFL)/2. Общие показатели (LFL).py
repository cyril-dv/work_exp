import numpy as np
import pandas as pd
np.set_printoptions(suppress=True, precision=2)
pd.options.display.float_format = '{:,.2f}'.format

import openpyxl as op
import duckdb

import os
import sys
from datetime import datetime


try:
    LOCAL_DIR = os.path.dirname(__file__)
except:
    LOCAL_DIR = os.getcwd()

PARENT_DIR = os.path.dirname(LOCAL_DIR)
sys.path.append(PARENT_DIR)

import tbl_calc

REPORT_MTH = tbl_calc.REPORT_MTH
DB_FILE = tbl_calc.DB_FILE


print('\n2. Общие показатели (LFL)')
print('-'*79)


groups_dict = {
    'Группа 1': 'a. Группа 1',
    'Группа 2': 'b. Группа 2',
    'Группа 3': 'c. Группа 3',
    'Группа 4': 'd. Группа 4',
    'Группа 5': 'e. Группа 5',
    'Группа 6': 'f. Группа 6'
}


try:
    conx = duckdb.connect(database=DB_FILE, read_only=True)
except Exception as err:
    print(f'Не удается подключиться к базе данных:\n{err}')


try:
    with open(os.path.join(LOCAL_DIR, '2. Общие показатели (LFL).sql'), encoding='utf-8') as q:
        sql_query = q.read()
except Exception as err:
    print(f'Не удается открыть файл с SQL запросом:\n{err}')


df = conx.execute(sql_query).fetch_df()
conx.close()


df.to_csv(os.path.join(LOCAL_DIR, 'data.csv'), index=False)


df['group'] = df['group'].replace(groups_dict)

df_tn = df.drop(columns=['thou_pcs', 'mil_rub'])
df_pcs = df.drop(columns=['tonnes', 'mil_rub'])
df_rub = df.drop(columns=['tonnes', 'thou_pcs'])


df_tn = pd.pivot(df_tn, columns='report_date', index='group', values='tonnes')
df_pcs = pd.pivot(df_pcs, columns='report_date', index='group', values='thou_pcs')
df_rub = pd.pivot(df_rub, columns='report_date', index='group', values='mil_rub')
df_price = (df_rub*1000)/df_tn


df_tn = df_tn.fillna(0).pipe(tbl_calc.margins)
df_pcs = df_pcs.fillna(0).pipe(tbl_calc.margins)
df_rub = df_rub.fillna(0).pipe(tbl_calc.margins)
df_price = df_price.fillna(0)


df_tn = (pd.concat([
        df_tn, 
        df_tn.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m',     calc_method='pct_change', aggfunc='sum').rename('m_on_m_pct'), 
        df_tn.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='ytd_on_ytd', calc_method='pct_change', aggfunc='sum').rename('ytd_on_ytd_pct'), 
        df_tn.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m',     calc_method='abs_change', aggfunc='sum').rename('m_on_m_abs'), 
        df_tn.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='ytd_on_ytd', calc_method='abs_change', aggfunc='sum').rename('ytd_on_ytd_abs')
    ], axis=1)
    .reset_index(names='group').replace(r'^\w\S\s', '', regex=True)
    .pipe(tbl_calc.insert_wsp, 4)
)

df_pcs = (pd.concat([
        df_pcs, 
        df_pcs.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m',     calc_method='pct_change', aggfunc='sum').rename('m_on_m_pct'), 
        df_pcs.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='ytd_on_ytd', calc_method='pct_change', aggfunc='sum').rename('ytd_on_ytd_pct'), 
        df_pcs.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m',     calc_method='abs_change', aggfunc='sum').rename('m_on_m_abs'), 
        df_pcs.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='ytd_on_ytd', calc_method='abs_change', aggfunc='sum').rename('ytd_on_ytd_abs')
    ], axis=1)
    .reset_index(names='group').replace(r'^\w\S\s', '', regex=True)
    .pipe(tbl_calc.insert_wsp, 4)
)


df_rub = (pd.concat([
        df_rub, 
        df_rub.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m',     calc_method='pct_change', aggfunc='sum').rename('m_on_m_pct'), 
        df_rub.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='ytd_on_ytd', calc_method='pct_change', aggfunc='sum').rename('ytd_on_ytd_pct'), 
        df_rub.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m',     calc_method='abs_change', aggfunc='sum').rename('m_on_m_abs'), 
        df_rub.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='ytd_on_ytd', calc_method='abs_change', aggfunc='sum').rename('ytd_on_ytd_abs')
    ], axis=1)
    .reset_index(names='group').replace(r'^\w\S\s', '', regex=True)
    .pipe(tbl_calc.insert_wsp, 4)
)


df_price = (pd.concat([
        df_price, 
        df_price.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m',     calc_method='pct_change', aggfunc='sum').rename('m_on_m_pct'), 
        df_price.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='ytd_on_ytd', calc_method='pct_change', aggfunc='mean').rename('ytd_on_ytd_pct'), 
        df_price.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='m_on_m',     calc_method='abs_change', aggfunc='sum').rename('m_on_m_abs'), 
        df_price.apply(tbl_calc.changes, axis=1, curr_month=REPORT_MTH, period='ytd_on_ytd', calc_method='abs_change', aggfunc='mean').rename('ytd_on_ytd_abs')
    ], axis=1)
    .reset_index(names='group').replace(r'^\w\S\s', '', regex=True)
    .pipe(tbl_calc.insert_wsp, 4)
)


with pd.ExcelWriter(os.path.join(LOCAL_DIR, 'output.xlsx')) as w:
    df_tn.to_excel(w, sheet_name='tn_LFL', index=False)
    df_pcs.to_excel(w, sheet_name='pcs_LFL', index=False)
    df_rub.to_excel(w, sheet_name='rub_LFL', index=False)
    df_price.to_excel(w, sheet_name='price_LFL', index=False)


col_widths = [20] + [11]*(df_tn.shape[1] - 1)

wb = op.load_workbook(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))
side = op.styles.Side(border_style=None)
no_border = op.styles.borders.Border(left=side, right=side, top=side, bottom=side)
font_txt = op.styles.Font(name='Arial', size=10)
font_head = op.styles.Font(name='Arial', size=10, bold=True)
for ws in wb:
    for col in ws.iter_cols():
        ws.column_dimensions[op.utils.get_column_letter(col[0].column)].width = col_widths[col[0].column - 1]
        for cell in col:
            if cell.column > 1 and cell.row > 1:
                if ws.title == 'price_LFL':
                    cell.number_format = '#,##0.0'
                else:
                    cell.number_format = '#,##0'
            if 1 < cell.column < (df_tn.shape[1] - 4) and cell.row == 1:
                cell.number_format = 'MMM YYYY'
            if cell.row == 1:
                cell.font = font_head
                cell.border = no_border
            else:
                cell.font = font_txt
wb.save(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))


print(f'{datetime.now().strftime("%H:%M:%S")}> Готово')