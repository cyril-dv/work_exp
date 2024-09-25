import numpy as np
import pandas as pd
np.set_printoptions(suppress=True, precision=2)
pd.options.display.float_format = '{:,.2f}'.format

import openpyxl as op
import duckdb

import os
import sys
from datetime import datetime


try:
    LOCAL_DIR = os.path.dirname(__file__)
except:
    LOCAL_DIR = os.getcwd()

PARENT_DIR = os.path.dirname(LOCAL_DIR)
sys.path.append(PARENT_DIR)

import tbl_calc

REPORT_MTH = tbl_calc.REPORT_MTH
DB_FILE = tbl_calc.DB_FILE


def pretty_change(row):
    try:
        if row["prior_3m_1"]/row["prior_3m_2"]-1 > 1.5:
            chn_tn = f'{row["prior_3m_1"]/row["prior_3m_2"]:.0f}x'
        elif row["prior_3m_1"]/row["prior_3m_2"]-1 > 0:
            chn_tn = f'+{row["prior_3m_1"]/row["prior_3m_2"]-1:.0%}'
        else: 
            chn_tn = f'{row["prior_3m_1"]/row["prior_3m_2"]-1:.0%}'
    except:
        chn_tn = '–'

    return chn_tn.replace('.', ',')


print('\n11. Производители в сетях II')
print('-'*79)


network_list = ['network_1', 'network_2', 'network_3', 'network_4', 'network_5', 'network_6', 'network_7', 'network_8', 'network_9', 'network_10']
groups_dict = {
        'Группа 1': 'a. Группа 1',
        'Группа 2': 'b. Группа 2',
        'Группа 3': 'c. Группа 3',
        'Группа 4': 'd. Группа 4',
        'Группа 5': 'e. Группа 5',
        'Группа 6': 'f. Группа 6'
}


try:
    conx = duckdb.connect(database=DB_FILE, read_only=True)
except Exception as err:
    print(f'Не удается подключиться к базе данных:\n{err}')


try:
    with open(os.path.join(LOCAL_DIR, '11. Производители в сетях II.sql'), encoding='utf-8') as q:
        sql_query = q.read()
except Exception as err:
    print(f'Не удается открыть файл с SQL запросом:\n{err}')


df = conx.execute(sql_query).fetch_df()
conx.close()


df.to_csv(os.path.join(LOCAL_DIR, 'data.csv'), index=False)


df = df.pivot(index=['network_subname', 'group', 'producer'], columns='month_year', values='tn').reset_index()
df.columns.name = None

df = df.fillna(0)
df['group'] = df['group'].replace(groups_dict)
df['tn'] = df['prior_3m_1'].apply(lambda x: f'{x:,.0f}'.replace(',', ' '))
df['chn_tn'] = df.apply(pretty_change, axis=1)
df['val'] = df['tn'] + ' (' + df['chn_tn'] + ')'
df = df.replace({'0 (-100%)':None, '0 (–)':None})


grouped = df.groupby("group")


writer = pd.ExcelWriter(os.path.join(LOCAL_DIR, 'output.xlsx'))

for name, group in grouped:
    df_sort = group[['producer', 'prior_3m_1']].groupby('producer').sum()
    df_margins = group[['network_subname', 'prior_3m_1']].groupby('network_subname', as_index=False).sum()
    df_margins = pd.DataFrame(
                        {k:v for (k, v) in zip(sorted(network_list), df_margins['prior_3m_1'])}, 
                        index=['Все производители'])
    
    group = group.pivot(index='producer', columns='network_subname', values='val')
    group = group.reindex(columns=network_list)
    
    group = group.merge(df_sort, on='producer')
    group = group.sort_values(by='prior_3m_1', ascending=False)
    df_others = group.loc[['Прочие'], :].copy()

    group = group.drop(index='Прочие')
    group = pd.concat([group, df_others, df_margins], axis=0)
    group = group.reset_index(names='Производитель')
    group = group.rename(columns={'prior_3m_1':'Все сети'})
    group.iat[-1, -1] = group['Все сети'].sum()

    group.to_excel(writer, sheet_name=name[3:], index=False)

writer.close()


# Базовое форматирование
col_widths = [26] + [12]*11

wb = op.load_workbook(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))
side = op.styles.Side(border_style=None)
no_border = op.styles.borders.Border(left=side, right=side, top=side, bottom=side)
font_txt = op.styles.Font(name='Arial', size=10)
font_head = op.styles.Font(name='Arial', size=10, bold=True)
for ws in wb:
    for col in ws.iter_cols():
        ws.column_dimensions[op.utils.get_column_letter(col[0].column)].width = col_widths[col[0].column - 1]
        for cell in col:
            if cell.row == 1 or cell.column == 1:
                cell.font = font_head
                cell.border = no_border
            elif (cell.row == ws.max_row) or (cell.column == ws.max_column):
                cell.number_format = '#,##0'
            cell.font = font_txt
wb.save(filename=os.path.join(LOCAL_DIR, 'output.xlsx'))


print(f'{datetime.now().strftime("%H:%M:%S")}> Готово')